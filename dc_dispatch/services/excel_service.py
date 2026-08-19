from __future__ import annotations

from collections import defaultdict
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import cint, flt
from frappe.utils.file_manager import get_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Protection,
)
from openpyxl.utils import get_column_letter

from dc_dispatch.services.allocation import (
    validate_related_sets,
)
from dc_dispatch.services.dispatch_matrix_service import (
    build_dispatch_matrix,
)
from dc_dispatch.services.run_service import (
    assert_calculation_inputs_unchanged,
    assert_stock_snapshot,
    validate_current_proposal,
)


PROPOSAL_HEADERS = [
    "Line ID",
    "Run ID",
    "Revision",
    "Source DC",
    "Store Warehouse",
    "Transit Warehouse",
    "Item Template",
    "Item Variant / Size",
    "Related Set",
    "Main Group",
    "Historical Demand Score",
    "Store Share %",
    "Suggested Qty",
    "Final Qty",
    "Exclude",
    "Override Reason",
    "Validation Status",
]

EDITABLE_HEADERS = {
    "Final Qty",
    "Exclude",
    "Override Reason",
}


@frappe.whitelist()
def download_proposal(run_name):
    run = frappe.get_doc(
        "DC Dispatch Run",
        run_name,
    )
    run.check_permission("read")
    return export_proposal(run)


def export_proposal(run):
    if run.status not in {
        "Calculated",
        "Proposal Imported",
    }:
        frappe.throw(
            _("Calculate the proposal before exporting it.")
        )

    assert_calculation_inputs_unchanged(run)

    lines = _proposal_lines(run)
    if not lines:
        frappe.throw(
            _("No proposal lines exist for this revision.")
        )

    workbook = Workbook()

    summary = workbook.active
    summary.title = "Run Summary"
    _write_summary(
        summary,
        run,
        lines,
    )

    _write_style_summary(
        workbook.create_sheet("Style Summary"),
        run,
    )

    _write_proposal(
        workbook.create_sheet("Allocation Proposal"),
        lines,
    )

    _write_simple_allocation(
        workbook.create_sheet("Simple Allocation"),
        run,
        lines,
    )

    _write_warnings(
        workbook.create_sheet("Warnings"),
        run,
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    frappe.local.response.filename = (
        f"{run.name}-R{run.revision}-DC-Dispatch.xlsx"
    )
    frappe.local.response.filecontent = output.read()
    frappe.local.response.type = "binary"


def import_proposal(run):
    if run.status not in {
        "Calculated",
        "Proposal Imported",
    }:
        frappe.throw(
            _("Only a calculated proposal can be imported.")
        )

    if not run.proposal_file:
        frappe.throw(
            _("Attach the reviewed proposal workbook first.")
        )

    assert_calculation_inputs_unchanged(run)
    assert_stock_snapshot(run)

    _filename, content = get_file(
        run.proposal_file
    )

    try:
        workbook = load_workbook(
            BytesIO(content),
            data_only=False,
            read_only=False,
        )
    except Exception as exc:
        frappe.throw(
            _(
                "The attached file is not a valid Excel "
                "workbook: {0}"
            ).format(exc)
        )

    if "Allocation Proposal" not in workbook.sheetnames:
        frappe.throw(
            _(
                "The workbook does not contain the "
                "Allocation Proposal sheet."
            )
        )

    sheet = workbook["Allocation Proposal"]

    headers = {
        cell.value: index + 1
        for index, cell in enumerate(sheet[1])
        if cell.value
    }

    missing_headers = [
        header
        for header in PROPOSAL_HEADERS
        if header not in headers
    ]

    if missing_headers:
        frappe.throw(
            _("Missing workbook columns: {0}").format(
                ", ".join(missing_headers)
            )
        )

    database_lines = {
        row.name: row
        for row in frappe.get_all(
            "DC Dispatch Proposal Line",
            filters={
                "run": run.name,
                "revision": run.revision,
            },
            fields=[
                "name",
                "run",
                "revision",
                "store_warehouse",
                "item_template",
                "item_code",
                "related_set",
                "suggested_qty",
                "final_qty",
                "exclude",
            ],
            limit_page_length=0,
        )
    }

    imported = {}
    errors = []

    for row_number in range(
        2,
        sheet.max_row + 1,
    ):
        line_id = sheet.cell(
            row_number,
            headers["Line ID"],
        ).value

        if (
            not line_id
            and all(
                sheet.cell(
                    row_number,
                    column,
                ).value in (None, "")
                for column in range(
                    1,
                    sheet.max_column + 1,
                )
            )
        ):
            continue

        if line_id not in database_lines:
            errors.append(
                f"Row {row_number}: unknown Line ID {line_id}."
            )
            continue

        if line_id in imported:
            errors.append(
                f"Row {row_number}: duplicate Line ID {line_id}."
            )
            continue

        run_id = sheet.cell(
            row_number,
            headers["Run ID"],
        ).value

        revision = sheet.cell(
            row_number,
            headers["Revision"],
        ).value

        if (
            run_id != run.name
            or cint(revision) != cint(run.revision)
        ):
            errors.append(
                f"Row {row_number}: run ID or revision "
                "does not match the current proposal."
            )
            continue

        final_value = sheet.cell(
            row_number,
            headers["Final Qty"],
        ).value

        try:
            final_number = float(
                final_value or 0
            )
        except (TypeError, ValueError):
            errors.append(
                f"Row {row_number}: Final Qty must "
                "be a whole number."
            )
            continue

        if (
            final_number < 0
            or not final_number.is_integer()
        ):
            errors.append(
                f"Row {row_number}: Final Qty must be "
                "a non-negative whole number."
            )
            continue

        exclude = _as_check(
            sheet.cell(
                row_number,
                headers["Exclude"],
            ).value
        )

        reason = str(
            sheet.cell(
                row_number,
                headers["Override Reason"],
            ).value
            or ""
        ).strip()

        original = database_lines[line_id]

        changed = (
            cint(final_number)
            != cint(original.suggested_qty)
            or exclude
        )

        if changed and not reason:
            errors.append(
                f"Row {row_number}: Override Reason is "
                "required when changing or excluding a line."
            )

        imported[line_id] = {
            "final_qty": cint(final_number),
            "exclude": exclude,
            "override_reason": reason,
        }

    missing_lines = (
        set(database_lines)
        - set(imported)
    )

    if missing_lines:
        errors.append(
            f"The workbook is missing "
            f"{len(missing_lines)} proposal lines."
        )

    if errors:
        frappe.throw(
            "<br>".join(errors[:50])
        )

    _validate_imported_totals(
        run,
        database_lines,
        imported,
    )

    for line_id, values in imported.items():
        frappe.db.set_value(
            "DC Dispatch Proposal Line",
            line_id,
            values,
            update_modified=False,
        )

    run.status = "Proposal Imported"
    run.save()

    validate_current_proposal(run)

    return {
        "lines": len(imported),
        "status": run.status,
    }


def _validate_imported_totals(
    run,
    lines,
    imported,
):
    snapshots = {
        row.item_code: flt(row.actual_qty)
        for row in frappe.get_all(
            "DC Dispatch Stock Snapshot",
            filters={
                "run": run.name,
                "revision": run.revision,
            },
            fields=[
                "item_code",
                "actual_qty",
            ],
            limit_page_length=0,
        )
    }

    variant_totals = defaultdict(int)
    template_totals = defaultdict(int)
    validation_rows = []

    for line_id, original in lines.items():
        values = imported[line_id]

        quantity = (
            0
            if values["exclude"]
            else values["final_qty"]
        )

        variant_totals[
            original.item_code
        ] += quantity

        template_totals[
            original.item_template
        ] += quantity

        validation_rows.append(
            {
                "store_warehouse": (
                    original.store_warehouse
                ),
                "item_template": (
                    original.item_template
                ),
                "related_set": (
                    original.related_set
                ),
                "final_qty": quantity,
                "exclude": values["exclude"],
            }
        )

    errors = []

    for item_code, quantity in (
        variant_totals.items()
    ):
        if quantity > snapshots.get(
            item_code,
            0,
        ):
            errors.append(
                f"{item_code}: final {quantity} exceeds "
                f"snapshot "
                f"{snapshots.get(item_code, 0):g}."
            )

    targets = {
        row.item_template: cint(row.target_qty)
        for row in run.items
    }

    for template, quantity in (
        template_totals.items()
    ):
        if quantity > targets.get(
            template,
            0,
        ):
            errors.append(
                f"{template}: final {quantity} exceeds "
                f"dispatch target "
                f"{targets.get(template, 0)}."
            )

    expected_members = defaultdict(set)

    for row in run.items:
        if row.related_set:
            expected_members[
                row.related_set
            ].add(row.item_template)

    errors.extend(
        validate_related_sets(
            validation_rows,
            expected_members,
        )
    )

    if errors:
        frappe.throw(
            "<br>".join(errors[:50])
        )


def _proposal_lines(run):
    return frappe.get_all(
        "DC Dispatch Proposal Line",
        filters={
            "run": run.name,
            "revision": run.revision,
        },
        fields=[
            "name",
            "run",
            "revision",
            "source_warehouse",
            "store_warehouse",
            "transit_warehouse",
            "item_template",
            "item_code",
            "related_set",
            "main_group",
            "sales_score",
            "share_percent",
            "suggested_qty",
            "final_qty",
            "exclude",
            "override_reason",
            "validation_status",
        ],
        order_by=(
            "item_template asc, "
            "store_warehouse asc, "
            "item_code asc"
        ),
        limit_page_length=0,
    )


def _write_summary(
    sheet,
    run,
    lines,
):
    rows = [
        ("DC Dispatch Run", run.name),
        ("Revision", run.revision),
        ("Company", run.company),
        ("Source DC", run.source_warehouse),
        (
            "Sales Period",
            (
                f"{run.sales_from_date} "
                f"to {run.sales_to_date}"
            ),
        ),
        ("Status at Export", run.status),
        ("Styles", len(run.items)),
        (
            "Eligible Stores",
            sum(
                1
                for row in run.store_rules
                if row.decision != "Exclude"
            ),
        ),
        (
            "Suggested Quantity",
            sum(
                cint(row.suggested_qty)
                for row in lines
            ),
        ),
    ]

    for row in rows:
        sheet.append(row)

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 48

    for cell in sheet[1]:
        cell.font = Font(bold=True)


def _write_style_summary(
    sheet,
    run,
):
    headers = [
        "Item Template",
        "Main Group",
        "Subgroup",
        "Related Set",
        "Available DC Qty",
        "Dispatch %",
        "Dispatch Target",
        "Matching Templates",
        "Cohort Net Units",
        "Stores with Sales",
        "Warning",
    ]

    sheet.append(headers)

    for row in run.items:
        sheet.append(
            [
                row.item_template,
                row.main_group,
                row.subgroup,
                row.related_set,
                row.dc_qty,
                row.dispatch_percentage,
                row.target_qty,
                row.cohort_templates,
                row.cohort_units,
                row.cohort_stores,
                row.warning,
            ]
        )

    _format_table(
        sheet,
        editable_columns=set(),
    )


def _write_proposal(
    sheet,
    lines,
):
    sheet.append(PROPOSAL_HEADERS)

    for row in lines:
        sheet.append(
            [
                row.name,
                row.run,
                row.revision,
                row.source_warehouse,
                row.store_warehouse,
                row.transit_warehouse,
                row.item_template,
                row.item_code,
                row.related_set,
                row.main_group,
                row.sales_score,
                row.share_percent,
                row.suggested_qty,
                row.final_qty,
                (
                    "Yes"
                    if row.exclude
                    else "No"
                ),
                row.override_reason,
                row.validation_status,
            ]
        )

    _format_table(
        sheet,
        editable_columns=EDITABLE_HEADERS,
    )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.protection.sheet = True
    sheet.protection.password = "dcdispatch"


def _write_simple_allocation(
    sheet,
    run,
    lines,
):
    matrix = build_dispatch_matrix(
        run,
        lines=lines,
    )

    headers = (
        ["Item Template", "Size"]
        + matrix["stores"]
        + [
            "Total Dispatched",
            "Total DC Qty",
            "Remaining Qty",
        ]
    )
    sheet.append(headers)

    for row in matrix["rows"]:
        sheet.append(
            [
                row["item_template"],
                row["size"],
                *[
                    row["store_quantities"].get(
                        store,
                        0,
                    )
                    for store in matrix["stores"]
                ],
                row["total_dispatched"],
                row["total_dc_qty"],
                row["remaining_qty"],
            ]
        )

    total_row = sheet.max_row + 1
    sheet.cell(
        total_row,
        1,
        "Total",
    )

    for column in range(
        3,
        sheet.max_column + 1,
    ):
        column_letter = get_column_letter(
            column
        )
        sheet.cell(
            total_row,
            column,
            (
                f"=SUM("
                f"{column_letter}2:"
                f"{column_letter}"
                f"{total_row - 1})"
            ),
        )

    header_fill = PatternFill(
        "solid",
        fgColor="1F7A35",
    )
    body_fill = PatternFill(
        "solid",
        fgColor="C6EFCE",
    )
    alternate_fill = PatternFill(
        "solid",
        fgColor="A9E6B2",
    )
    total_fill = PatternFill(
        "solid",
        fgColor="1F7A35",
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row_number in range(
        2,
        total_row,
    ):
        fill = (
            body_fill
            if row_number % 2 == 0
            else alternate_fill
        )

        for cell in sheet[row_number]:
            cell.fill = fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

    for cell in sheet[total_row]:
        cell.fill = total_fill
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    sheet.freeze_panes = "C2"

    if total_row > 2:
        sheet.auto_filter.ref = (
            f"A1:"
            f"{get_column_letter(sheet.max_column)}"
            f"{total_row - 1}"
        )

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 10

    first_store_col = 3
    last_store_col = (
        first_store_col
        + len(matrix["stores"])
        - 1
    )

    for column in range(
        first_store_col,
        last_store_col + 1,
    ):
        sheet.column_dimensions[
            get_column_letter(column)
        ].width = 18

    for column in range(
        last_store_col + 1,
        sheet.max_column + 1,
    ):
        sheet.column_dimensions[
            get_column_letter(column)
        ].width = 16


def _write_warnings(
    sheet,
    run,
):
    sheet.append(
        [
            "Type",
            "Record",
            "Warning / Decision",
        ]
    )

    for row in run.items:
        if row.warning:
            sheet.append(
                [
                    "Weak Historical Evidence",
                    row.item_template,
                    row.warning,
                ]
            )

    for row in run.store_rules:
        if row.history_status == "No History":
            decision = row.decision

            if row.reference_store:
                decision += (
                    f": {row.reference_store}"
                )

            sheet.append(
                [
                    "Store Without History",
                    row.store_warehouse,
                    decision,
                ]
            )

    _format_table(
        sheet,
        editable_columns=set(),
    )


def _format_table(
    sheet,
    editable_columns,
):
    header_fill = PatternFill(
        "solid",
        fgColor="551C25",
    )
    editable_fill = PatternFill(
        "solid",
        fgColor="FFF2CC",
    )

    header_map = {
        cell.value: cell.column
        for cell in sheet[1]
    }

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for header, column in (
        header_map.items()
    ):
        width = min(
            45,
            max(
                12,
                len(str(header)) + 2,
            ),
        )

        for cell in sheet[
            get_column_letter(column)
        ]:
            if cell.row > 1:
                width = min(
                    45,
                    max(
                        width,
                        len(
                            str(
                                cell.value or ""
                            )
                        )
                        + 2,
                    ),
                )

                if header in editable_columns:
                    cell.protection = (
                        Protection(
                            locked=False
                        )
                    )
                    cell.fill = editable_fill
                else:
                    cell.protection = (
                        Protection(
                            locked=True
                        )
                    )

        sheet.column_dimensions[
            get_column_letter(column)
        ].width = width


def _as_check(value):
    if isinstance(value, bool):
        return int(value)

    if isinstance(
        value,
        (int, float),
    ):
        return int(bool(value))

    return int(
        str(value or "")
        .strip()
        .casefold()
        in {
            "yes",
            "y",
            "true",
            "1",
            "exclude",
        }
    )
