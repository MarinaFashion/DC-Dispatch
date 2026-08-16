from __future__ import annotations

from collections import defaultdict
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import flt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dc_dispatch.services.run_service import (
    _adjust_store_scores,
    _fields_by_main_group,
    _has_value,
    _historical_sales,
    _item_values,
    _normal,
    _reference_fieldnames,
    assert_calculation_inputs_unchanged,
)


@frappe.whitelist()
def download_history_evidence(run_name):
    """Export the exact aggregated historical evidence used by DC Dispatch.

    The workbook is an audit/reconciliation aid. It uses the same sales-period,
    return/store treatment, matching threshold, included stores, and reference
    store score substitution as proposal calculation.
    """
    run = frappe.get_doc("DC Dispatch Run", run_name)
    run.check_permission("read")
    return export_history_evidence(run)


def export_history_evidence(run):
    if not run.items:
        frappe.throw(_("Load target items before exporting historical evidence."))
    if not run.reference_fields:
        frappe.throw(_("Select historical matching fields before exporting historical evidence."))
    if not run.store_rules:
        frappe.throw(_("Load eligible stores before exporting historical evidence."))

    # If the run already has a calculated revision, ensure this audit is for the
    # same immutable criteria that generated that revision.
    if run.calculation_input_hash:
        assert_calculation_inputs_unchanged(run)

    settings = frappe.get_single("DC Dispatch Settings")
    active_rules = [row for row in run.store_rules if row.decision != "Exclude"]
    stores = [row.store_warehouse for row in active_rules]
    if not stores:
        frappe.throw(_("No included stores exist in this run."))

    sales = _historical_sales(run, stores)
    candidate_templates = {template for template, _store in sales}
    target_templates = {row.item_template for row in run.items}

    value_fields = _reference_fieldnames(run) | {
        settings.item_main_group_field,
        settings.item_subgroup_field,
        settings.item_related_set_field,
    }
    values = _item_values(candidate_templates | target_templates, value_fields)
    fields_by_group = _fields_by_main_group(run)
    threshold = flt(run.minimum_match_percent)

    field_labels = _field_labels(value_fields)
    proposal_scores = _current_proposal_scores(run)

    target_summaries = []
    cohort_rows = []
    sales_rows = []
    score_rows = []

    for item_row in run.items:
        target_values = values.get(item_row.item_template, {})
        group = target_values.get(settings.item_main_group_field) or item_row.main_group
        selected_fields = fields_by_group.get(group, [])
        comparable_fields = [
            field for field in selected_fields if _has_value(target_values.get(field))
        ]

        cohort = []
        match_details = {}

        for candidate in sorted(candidate_templates):
            candidate_values = values.get(candidate, {})
            if candidate_values.get(settings.item_main_group_field) != group:
                continue

            if not comparable_fields:
                match_percent = 100.0
            else:
                matches = sum(
                    1
                    for field in comparable_fields
                    if _has_value(candidate_values.get(field))
                    and _normal(candidate_values.get(field)) == _normal(target_values.get(field))
                )
                match_percent = matches * 100.0 / len(comparable_fields)

            if match_percent < threshold:
                continue

            cohort.append(candidate)
            details = []
            for field in selected_fields:
                target_value = target_values.get(field)
                historical_value = candidate_values.get(field)
                comparable = _has_value(target_value)
                matched = (
                    comparable
                    and _has_value(historical_value)
                    and _normal(historical_value) == _normal(target_value)
                )
                details.append(
                    {
                        "field": field,
                        "label": field_labels.get(field, field),
                        "target": target_value,
                        "historical": historical_value,
                        "comparable": comparable,
                        "matched": matched,
                    }
                )
            match_details[candidate] = {
                "match_percent": match_percent,
                "details": details,
            }

        raw_scores = defaultdict(float)
        cohort_set = set(cohort)
        for (template, store), quantity in sales.items():
            if template in cohort_set:
                raw_scores[store] += quantity
                if quantity:
                    sales_rows.append(
                        {
                            "target": item_row.item_template,
                            "historical": template,
                            "store": store,
                            "net_qty": flt(quantity),
                        }
                    )

        raw_scores = {store: max(0, flt(quantity)) for store, quantity in raw_scores.items()}
        adjusted_scores, missing_references = _adjust_store_scores(run, raw_scores)
        adjusted_total = sum(max(0, flt(value)) for value in adjusted_scores.values())

        target_summaries.append(
            {
                "target": item_row.item_template,
                "main_group": item_row.main_group,
                "subgroup": item_row.subgroup,
                "selected_fields": ", ".join(field_labels.get(field, field) for field in selected_fields),
                "comparable_fields": ", ".join(field_labels.get(field, field) for field in comparable_fields),
                "threshold": threshold,
                "cohort_templates": len(cohort),
                "cohort_units": sum(raw_scores.values()),
                "cohort_stores": sum(1 for value in raw_scores.values() if value > 0),
                "missing_references": ", ".join(missing_references),
            }
        )

        for candidate in cohort:
            detail = match_details[candidate]
            cohort_rows.append(
                {
                    "target": item_row.item_template,
                    "historical": candidate,
                    "match_percent": detail["match_percent"],
                    "match_detail": _match_detail_text(detail["details"]),
                    "historical_net_units": sum(
                        flt(quantity)
                        for (template, _store), quantity in sales.items()
                        if template == candidate
                    ),
                }
            )

        rules_by_store = {row.store_warehouse: row for row in active_rules}
        for store, rule in rules_by_store.items():
            raw = max(0, flt(raw_scores.get(store, 0)))
            applied = max(0, flt(adjusted_scores.get(store, 0)))
            score_rows.append(
                {
                    "target": item_row.item_template,
                    "store": store,
                    "decision": rule.decision,
                    "reference_store": rule.reference_store,
                    "raw_score": raw,
                    "applied_score": applied,
                    "share_percent": applied * 100.0 / adjusted_total if adjusted_total else 0,
                    "proposal_score": proposal_scores.get((item_row.item_template, store)),
                }
            )

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Run Summary"
    _write_run_summary(summary, run, active_rules)

    _write_target_summary(workbook.create_sheet("Target Cohorts"), target_summaries)
    _write_cohort_templates(workbook.create_sheet("Historical Templates"), cohort_rows)
    _write_sales_used(workbook.create_sheet("Sales by Store"), sales_rows)
    _write_score_reconciliation(workbook.create_sheet("Score Reconciliation"), score_rows)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    frappe.local.response.filename = (
        f"{run.name}-Historical-Evidence"
        + (f"-R{run.revision}" if run.revision else "")
        + ".xlsx"
    )
    frappe.local.response.filecontent = output.read()
    frappe.local.response.type = "binary"


def _current_proposal_scores(run):
    if not run.revision:
        return {}
    rows = frappe.get_all(
        "DC Dispatch Proposal Line",
        filters={"run": run.name, "revision": run.revision},
        fields=["item_template", "store_warehouse", "sales_score"],
        order_by="item_template asc, store_warehouse asc",
        limit_page_length=0,
    )
    result = {}
    for row in rows:
        key = (row.item_template, row.store_warehouse)
        if key not in result:
            result[key] = flt(row.sales_score)
    return result


def _field_labels(fieldnames):
    meta = frappe.get_meta("Item")
    result = {}
    for fieldname in fieldnames:
        if not fieldname:
            continue
        field = meta.get_field(fieldname)
        result[fieldname] = field.label if field and field.label else fieldname
    return result


def _match_detail_text(details):
    parts = []
    for detail in details:
        if not detail["comparable"]:
            result = "Target blank / ignored"
        elif detail["matched"]:
            result = "Match"
        else:
            result = "No match"
        parts.append(
            f'{detail["label"]}: target="{detail["target"] or ""}", '
            f'history="{detail["historical"] or ""}" [{result}]'
        )
    return " | ".join(parts)


def _write_run_summary(sheet, run, active_rules):
    rows = [
        ("DC Dispatch Run", run.name),
        ("Revision", run.revision or 0),
        ("Status", run.status),
        ("Company", run.company),
        ("Source DC", run.source_warehouse),
        ("Sales From Date", run.sales_from_date),
        ("Sales To Date", run.sales_to_date),
        ("Minimum Field Match %", run.minimum_match_percent),
        ("Target Styles", len(run.items)),
        ("Included Stores", len(active_rules)),
        (
            "Matching Configuration",
            " ; ".join(
                f"{row.main_group}: {row.field_label or row.fieldname}"
                for row in run.reference_fields
            ),
        ),
        (
            "Return Warehouse Rule",
            "Linked returns use original Sales Invoice Item warehouse; otherwise return-line warehouse.",
        ),
        (
            "Evidence Granularity",
            "Net units aggregated by historical Item Template and Store Warehouse.",
        ),
    ]
    sheet.append(["Parameter", "Value"])
    for row in rows:
        sheet.append(row)
    _format_sheet(sheet, freeze="A2", auto_filter=True)


def _write_target_summary(sheet, rows):
    headers = [
        "Target Item Template",
        "Main Group",
        "Subgroup",
        "Selected Matching Fields",
        "Comparable Target Fields",
        "Required Match %",
        "Matching Historical Templates",
        "Cohort Net Units",
        "Stores with Positive Cohort Sales",
        "Reference Store Warning",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                row["target"],
                row["main_group"],
                row["subgroup"],
                row["selected_fields"],
                row["comparable_fields"],
                row["threshold"],
                row["cohort_templates"],
                row["cohort_units"],
                row["cohort_stores"],
                row["missing_references"],
            ]
        )
    _format_sheet(sheet, freeze="A2", auto_filter=True)


def _write_cohort_templates(sheet, rows):
    headers = [
        "Target Item Template",
        "Historical Item Template Used",
        "Actual Match %",
        "Match Detail",
        "Historical Template Net Units in Selected Stores/Period",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                row["target"],
                row["historical"],
                row["match_percent"],
                row["match_detail"],
                row["historical_net_units"],
            ]
        )
    _format_sheet(sheet, freeze="A2", auto_filter=True)


def _write_sales_used(sheet, rows):
    headers = [
        "Target Item Template",
        "Historical Item Template Used",
        "Store Warehouse",
        "Net Units Used",
    ]
    sheet.append(headers)
    for row in sorted(
        rows,
        key=lambda value: (
            value["target"],
            value["historical"],
            value["store"],
        ),
    ):
        sheet.append(
            [
                row["target"],
                row["historical"],
                row["store"],
                row["net_qty"],
            ]
        )
    _format_sheet(sheet, freeze="A2", auto_filter=True)


def _write_score_reconciliation(sheet, rows):
    headers = [
        "Target Item Template",
        "Final Store",
        "Decision",
        "Reference Store",
        "Raw Cohort Score at Final Store",
        "Applied Historical Demand Score",
        "Applied Store Share %",
        "Score Stored on Current Proposal",
        "Reconciles to Proposal",
    ]
    sheet.append(headers)
    for row in rows:
        proposal_score = row["proposal_score"]
        reconciles = (
            ""
            if proposal_score is None
            else "Yes"
            if abs(flt(proposal_score) - flt(row["applied_score"])) < 0.000001
            else "No"
        )
        sheet.append(
            [
                row["target"],
                row["store"],
                row["decision"],
                row["reference_store"],
                row["raw_score"],
                row["applied_score"],
                row["share_percent"],
                proposal_score,
                reconciles,
            ]
        )
    _format_sheet(sheet, freeze="A2", auto_filter=True)


def _format_sheet(sheet, freeze=None, auto_filter=False):
    header_fill = PatternFill("solid", fgColor="551C25")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column in range(1, sheet.max_column + 1):
        width = 12
        for cell in sheet[get_column_letter(column)]:
            width = min(70, max(width, len(str(cell.value or "")) + 2))
            if cell.row > 1:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.column_dimensions[get_column_letter(column)].width = width

    if freeze:
        sheet.freeze_panes = freeze
    if auto_filter and sheet.max_row >= 1:
        sheet.auto_filter.ref = sheet.dimensions
